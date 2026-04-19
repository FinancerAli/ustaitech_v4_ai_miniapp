"""Export orders to Excel using openpyxl."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


async def generate_orders_excel(orders) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    headers = ["#", "Sana", "Foydalanuvchi", "Xizmat", "Narx", "Chegirma%", "Yakuniy narx", "Kupon", "Status", "Izoh"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_colors = {
        "pending": "FFF2CC",
        "confirmed": "E2EFDA",
        "rejected": "FCE4D6",
        "cancelled": "EDEDED",
    }

    for row_idx, order in enumerate(orders, 2):
        uname = order["username"] or ""
        full_name = order["full_name"] or ""
        user_label = f"{full_name} (@{uname})" if uname else full_name

        ws.cell(row=row_idx, column=1, value=order["id"])
        ws.cell(row=row_idx, column=2, value=str(order["created_at"])[:16])
        ws.cell(row=row_idx, column=3, value=user_label)
        ws.cell(row=row_idx, column=4, value=order["service_name"])
        ws.cell(row=row_idx, column=5, value=order["price"])
        ws.cell(row=row_idx, column=6, value=order["discount"] or 0)
        ws.cell(row=row_idx, column=7, value=order["final_price"] or order["price"])
        ws.cell(row=row_idx, column=8, value=order["coupon_code"] or "")
        ws.cell(row=row_idx, column=9, value=order["status"])
        ws.cell(row=row_idx, column=10, value=order["note"] or "")

        fill_color = status_colors.get(order["status"], "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).fill = row_fill

    col_widths = [6, 17, 25, 20, 12, 12, 15, 12, 12, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
